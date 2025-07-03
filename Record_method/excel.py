'''
TODO : 
- copy past to make a SQL base function
- add a function to copy past the actual more efficiently
'''

import openpyxl
import logging
from os import makedirs
from datetime import datetime
from threading import Thread, Event
from time import sleep
import tkinter as tk
from tkinter import messagebox

logging.basicConfig(filename='log\\trace.log',level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class Excel_recorder:
    def __init__(self,config,data):
        self.path = config["recording"]["Excel"]["FOLDER_PATH"]
        self.record_time = config["recording"]["Excel"]["RECORD_TIME"]
        self.data_structure = [d[0] for d in config["data_structure"]]
        self.full_path = f"{self.path}\\Recording_{datetime.now().date()}.xlsx"
        self.header = ["RECORD_TIME", "ROUTE_NAME", "ADS_TIME", *self.data_structure]
        self.data = data
        makedirs(self.path, exist_ok=True)
        
        self.stop_event = Event()
        self.excel_recorder = Thread(target=self.write_to_excel, args = [self.stop_event])

    def start(self):
        self.Excel_Open = self.initialize_excel()
        if not self.excel_recorder.is_alive() and self.Excel_Open:
            self.excel_recorder.start()

    def stop(self):
        if self.Excel_Open:
            self.stop_event.set()
            self.excel_recorder.join()

    def initialize_excel(self):
        try:
            self.wb = openpyxl.load_workbook(self.full_path)
            logging.info("Opening existing workbook")
            self.sheet = self.wb.worksheets[0]
            return True
        except FileNotFoundError:
            self.wb =  openpyxl.Workbook()
            logging.info("Creating new workbook")
            self.sheet = self.wb.worksheets[0]
            for col, value in enumerate(self.header, start=1):
                self.sheet.title = "data"
                self.sheet.cell(row=1, column=col, value=value)
            self.save_workbook(self.wb)
            return True
        except PermissionError:
            logging.info("Permission denied : Creating a new workbook")
            self.full_path= f"{self.path}\\Recording_{datetime.now()}.xlsx"
            self.wb =  openpyxl.Workbook()
            self.sheet = self.wb.worksheets[0]
            for col, value in enumerate(self.header, start=1):
                self.sheet.title = "data"
                self.sheet.cell(row=1, column=col, value=value)
            self.save_workbook(self.wb)
            return True
        except Exception as e:
            logging.info("Error while trying to open the file :",str(e))
            return False

    def save_workbook(self,wb:openpyxl.Workbook):
        logging.info(f"Saving workbook {self.full_path}")
        try:
            wb.save(self.full_path)
        except PermissionError as e:
            logging.error(f"Error while saving the excel file : {e}")

    def write_to_excel(self, stop_event):
        logging.info(f"Starting recording")
        while not stop_event.is_set():
            try:
                while not self.data.empty():
                    data=self.data.get()
                    data_value = [data['DATA'][d] for d in self.data_structure]
                    data_ordered = [datetime.now(),data["NAME"],data["DATE TIME"],*data_value]
                    self.sheet.append(data_ordered)
            except PermissionError:
                root = tk.Tk()
                root.withdraw()
                messagebox.showerror("Error", f"Permission error, please close the file : {self.full_path}")
            except Exception as e:
                logging.error(f"Error while writting into the excel file : {e}")
            self.save_workbook(self.wb)
            sleep(self.record_time)