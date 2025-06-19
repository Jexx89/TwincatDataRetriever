'''
TODO : 
- add a thread to record the data every second and limit the recording if the dataqueue in empty
- compelete the header to have all the data needed
- copy past to make a SQL base function
- add a function to copy past the actual more efficiently


'''

import openpyxl
import logging
from os import makedirs

PATH_TO_RECORDING = "recording/"
logging.basicConfig(filename='log\\trace.log',level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class Excel_recording:
    def __init__(self,route_cfg):
        self.route_cfg=route_cfg
        makedirs(PATH_TO_RECORDING, exist_ok=True)
        self.route_cfg["EXCEL_PATH"] = f"{PATH_TO_RECORDING}Recording_{self.route_cfg['ROUTE_NAME']}.xlsx"
        self.initialize_excel()

    def excel_record(self, data):
        self.sheet.append(data)

    def initialize_excel(self):
        try:
            workbook = openpyxl.load_workbook(self.route_cfg["EXCEL_PATH"])
            logging.info("Opening existing workbook")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            logging.info("Creating new workbook")
        sheet = workbook.active
        header = ["Timestamp", "Temperature"]
        for col, value in enumerate(header, start=1):
            sheet.cell(row=1, column=col, value=value)

        logging.info("Initialized workbook")
        self.workbook =  workbook
        self.sheet = sheet
        self.save_workbook()

    def save_workbook(self):
        logging.info(f"Saving workbook {self.route_cfg["EXCEL_PATH"]}")
        self.workbook.save(self.route_cfg["EXCEL_PATH"])

    def close_workbook(self):
        logging.info(f"Closing workbook {self.route_cfg["EXCEL_PATH"]}")
        self.save_workbook()
        self.workbook.close()
        self.sheet = None
        self.workbook= None

